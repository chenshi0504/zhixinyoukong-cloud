from __future__ import annotations

import csv
import importlib.util
import io
import json
import os
import re
import sys
import uuid
import zipfile
from collections import defaultdict
from functools import lru_cache
from pathlib import Path
from typing import Any
from urllib.parse import quote
from xml.etree import ElementTree

import httpx
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..config import get_settings
from ..database import get_db
from ..deps import get_current_user, require_role
from ..models.classroom import Class, student_classes
from ..models.self_algorithm import SelfAlgorithmSubmission
from ..models.user import User
from ..schemas.common import PagedResponse
from ..schemas.self_algorithm import SelfAlgorithmPack, SelfAlgorithmSubmissionRead

router = APIRouter(prefix="/api/cloud/self-algorithms", tags=["自研算法"])


_PROJECT_ROOT = Path(__file__).resolve().parents[4]
_PACK_ROOT = _PROJECT_ROOT / "backend" / "algorithms" / "student_packs"
_SUITE_MODULE_PATH = _PROJECT_ROOT / "backend" / "algorithms" / "student_algorithm_suite.py"

_ALGORITHM_PACKS = {
    "data_fusion": {
        "dir": "data_fusion",
        "label": "多源数据融合自研算法",
        "description": "基于路侧检测与点云特征，提交自研多源感知融合结果并与平台贝叶斯融合 baseline 对比。",
    },
    "traffic_reconstruction": {
        "dir": "traffic_reconstruction",
        "label": "交通流重构自研算法",
        "description": "基于 DTALite 路网、观测流量和验证样本，提交自研交通流重构结果。",
    },
    "traffic_prediction": {
        "dir": "traffic_prediction",
        "label": "交通流预测自研算法",
        "description": "基于历史交通流特征，提交短时流量预测结果并与平台决策树 baseline 对比。",
    },
    "signal_optimization": {
        "dir": "signal_optimization",
        "label": "上层-交通信号优化",
        "description": "基于单点交叉口场景，提交自研信号配时结果并与平台 baseline 对比。",
    },
    "path_guidance": {
        "dir": "path_guidance",
        "label": "中层-路径诱导",
        "description": "基于 SiouxFalls 路网 OD 与边权数据，提交自研路径诱导结果。",
    },
    "vehicle_speed_control": {
        "dir": "vehicle_speed_control",
        "label": "下层-车速控制",
        "description": "基于车辆控制指标模板，提交自研车速控制仿真结果。",
    },
    "coordination_control": {
        "dir": "coordination_control",
        "label": "三层级协同对比",
        "description": "基于无协同、局部协同、全协同 baseline，提交三层级协同控制对比结果。",
    },
}

_CODE_EXTS = {".py"}
_SPEC_EXTS = {".md", ".txt", ".docx"}
_RESULT_EXTS = {".xlsx", ".xls", ".csv", ".json", ".xml", ".docx", ".txt"}
_MAX_FILE_SIZE = 80 * 1024 * 1024


def _safe_name(filename: str | None, fallback: str) -> str:
    raw = Path(filename or fallback).name
    raw = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", raw).strip(". ")
    return raw or fallback


def _require_pack(algorithm_type: str) -> dict[str, Any]:
    item = _ALGORITHM_PACKS.get(algorithm_type)
    if not item:
        raise HTTPException(status_code=404, detail="自研算法类型不存在")
    pack_dir = _PACK_ROOT / item["dir"]
    if not pack_dir.is_dir():
        raise HTTPException(status_code=404, detail="基础数据包不存在")
    return item


def _read_metadata(pack_dir: Path) -> dict[str, Any]:
    metadata_path = pack_dir / "metadata.json"
    if not metadata_path.is_file():
        return {}
    try:
        return json.loads(metadata_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _pack_to_schema(algorithm_type: str, item: dict[str, Any]) -> SelfAlgorithmPack:
    pack_dir = _PACK_ROOT / item["dir"]
    metadata = _read_metadata(pack_dir)
    baseline_file = metadata.get("baseline_file")
    if not baseline_file and (pack_dir / "baseline_predictions.csv").is_file():
        baseline_file = "baseline_predictions.csv"
    primary_metric = metadata.get("primary_metric")
    if not primary_metric and isinstance(metadata.get("evaluation"), dict):
        primary_metric = metadata["evaluation"].get("primary_metric")
    return SelfAlgorithmPack(
        algorithm_type=algorithm_type,
        label=item["label"],
        description=item["description"],
        baseline_file=baseline_file,
        primary_metric=primary_metric,
        input_files=metadata.get("input_files") or [],
        submission_columns=metadata.get("submission_columns") or [],
        download_name=f"{algorithm_type}_基础数据包.zip",
    )


def _validate_upload(file: UploadFile, allowed_exts: set[str], label: str) -> tuple[str, bytes]:
    filename = _safe_name(file.filename, f"{label}.bin")
    ext = Path(filename).suffix.lower()
    if ext not in allowed_exts:
        raise HTTPException(
            status_code=400,
            detail=f"{label}文件类型不正确，允许扩展名：{', '.join(sorted(allowed_exts))}",
        )
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail=f"{label}文件为空")
    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"{label}文件超过 80MB 限制")
    return filename, content


def _get_student_class_id(db: Session, current_user: User, class_id: int | None) -> int:
    class_ids = db.execute(
        select(student_classes.c.class_id).where(student_classes.c.student_id == current_user.id)
    ).scalars().all()
    if class_id is not None:
        if class_id not in class_ids:
            raise HTTPException(status_code=403, detail="该班级不属于当前学生")
        return class_id
    if not class_ids:
        raise HTTPException(status_code=400, detail="当前账号尚未加入班级，无法提交自研算法")
    return int(class_ids[0])


def _save_file(folder: Path, filename: str, content: bytes) -> str:
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / filename
    path.write_bytes(content)
    return str(path)


def _download_headers(filename: str, ascii_fallback: str = "download.bin") -> dict[str, str]:
    encoded = quote(filename)
    return {"Content-Disposition": f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded}"}


def _try_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _avg_dict(rows: list[dict[str, Any]]) -> dict[str, float]:
    buckets: dict[str, list[float]] = defaultdict(list)
    for row in rows:
        for key, value in row.items():
            number = _try_float(value)
            if number is not None:
                buckets[str(key)].append(number)
    return {key: sum(values) / len(values) for key, values in buckets.items() if values}


def _extract_csv_metrics(path: Path) -> dict[str, float]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows = [row for row in reader if row]
    if rows and reader.fieldnames:
        metrics = _avg_dict(rows)
        if metrics:
            return metrics
    return _extract_text_metrics(text)


def _walk_json_numbers(value: Any, prefix: str = "") -> dict[str, list[float]]:
    found: dict[str, list[float]] = defaultdict(list)
    if isinstance(value, dict):
        for key, item in value.items():
            child_prefix = str(key) if not prefix else f"{prefix}.{key}"
            for child_key, numbers in _walk_json_numbers(item, child_prefix).items():
                found[child_key].extend(numbers)
    elif isinstance(value, list):
        for item in value:
            for child_key, numbers in _walk_json_numbers(item, prefix).items():
                found[child_key].extend(numbers)
    else:
        number = _try_float(value)
        if number is not None and prefix:
            found[prefix.split(".")[-1]].append(number)
    return found


def _extract_json_metrics(path: Path) -> dict[str, float]:
    data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
    buckets = _walk_json_numbers(data)
    return {key: sum(values) / len(values) for key, values in buckets.items() if values}


def _extract_xml_metrics(path: Path) -> dict[str, float]:
    root = ElementTree.parse(path).getroot()
    buckets: dict[str, list[float]] = defaultdict(list)
    for elem in root.iter():
        number = _try_float(elem.text)
        if number is not None:
            buckets[elem.tag.split("}")[-1]].append(number)
    return {key: sum(values) / len(values) for key, values in buckets.items() if values}


def _extract_xlsx_metrics(path: Path) -> dict[str, float]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        headers = [str(h).strip() if h is not None else f"col_{idx}" for idx, h in enumerate(rows[0])]
        dict_rows = []
        for row in rows[1:]:
            dict_rows.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
        return _avg_dict(dict_rows)
    except Exception:
        return {}


def _extract_docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            raw = zf.read("word/document.xml").decode("utf-8", errors="ignore")
            return re.sub(r"<[^>]+>", " ", raw)
    except Exception:
        return ""


def _extract_text_metrics(text: str) -> dict[str, float]:
    metrics: dict[str, float] = {}
    for key, value in re.findall(r"([A-Za-z_][\w.-]*)\s*[:=]\s*(-?\d+(?:\.\d+)?)", text):
        metrics[key] = float(value)
    if metrics:
        return metrics
    numbers = [float(x) for x in re.findall(r"-?\d+(?:\.\d+)?", text)]
    if numbers:
        return {"numeric_mean": sum(numbers) / len(numbers), "numeric_count": float(len(numbers))}
    return {}


def _extract_metrics(path: Path) -> dict[str, float]:
    suffix = path.suffix.lower()
    try:
        if suffix in {".csv", ".txt"}:
            return _extract_csv_metrics(path)
        if suffix == ".json":
            return _extract_json_metrics(path)
        if suffix == ".xml":
            return _extract_xml_metrics(path)
        if suffix in {".xlsx", ".xls"}:
            return _extract_xlsx_metrics(path)
        if suffix == ".docx":
            return _extract_text_metrics(_extract_docx_text(path))
    except Exception:
        return {}
    return {}


def _metric_prefers_lower(name: str) -> bool:
    lowered = name.lower()
    lower_terms = ("delay", "queue", "waiting", "cost", "rmse", "mae", "error", "loss", "time", "std")
    higher_terms = ("speed", "throughput", "flow", "capacity", "score", "green")
    if any(term in lowered for term in lower_terms):
        return True
    if any(term in lowered for term in higher_terms):
        return False
    return True


def _trim_dict(data: dict[str, float], limit: int = 20) -> dict[str, float]:
    items = list(data.items())[:limit]
    return {key: round(value, 6) for key, value in items}


@lru_cache(maxsize=1)
def _load_student_suite():
    if not _SUITE_MODULE_PATH.is_file():
        return None
    try:
        spec = importlib.util.spec_from_file_location("student_algorithm_suite", _SUITE_MODULE_PATH)
        if spec is None or spec.loader is None:
            return None
        module = importlib.util.module_from_spec(spec)
        sys.modules[spec.name] = module
        spec.loader.exec_module(module)
        return module
    except Exception:
        return None


def _run_suite_evaluation(algorithm_type: str, result_path: Path) -> tuple[dict[str, Any] | None, str | None]:
    module = _load_student_suite()
    if module is None or not hasattr(module, "evaluate"):
        return None, None
    try:
        result = module.evaluate(algorithm_type, result_path, _PACK_ROOT)
        return result if isinstance(result, dict) else None, None
    except Exception as exc:
        return None, str(exc)


def _flatten_numeric_metrics(data: Any, prefix: str = "") -> dict[str, float]:
    metrics: dict[str, float] = {}
    if not isinstance(data, dict):
        return metrics
    for key, value in data.items():
        name = str(key) if not prefix else f"{prefix}.{key}"
        if isinstance(value, dict):
            metrics.update(_flatten_numeric_metrics(value, name))
            continue
        number = _try_float(value)
        if number is not None:
            metrics[name] = number
    return metrics


def _suite_prefers_lower(metric: str, metadata: dict[str, Any]) -> bool:
    evaluation = metadata.get("evaluation") if isinstance(metadata.get("evaluation"), dict) else {}
    leaf = metric.split(".")[-1]
    if leaf in set(evaluation.get("smaller_is_better") or []):
        return True
    if leaf in set(evaluation.get("larger_is_better") or []):
        return False
    return _metric_prefers_lower(metric)


def _analysis_from_suite_evaluation(
    algorithm_type: str,
    metadata: dict[str, Any],
    evaluation: dict[str, Any],
) -> dict[str, Any]:
    task_name = metadata.get("task_name") or _ALGORITHM_PACKS.get(algorithm_type, {}).get("label") or algorithm_type
    student_metrics = _flatten_numeric_metrics(evaluation.get("student_metrics"))
    baseline_metrics = _flatten_numeric_metrics(evaluation.get("baseline_metrics"))
    if not student_metrics:
        student_metrics = _flatten_numeric_metrics(evaluation.get("reference_metrics"))
    comparison_metrics = _flatten_numeric_metrics(evaluation.get("comparison"))

    comparisons = []
    advantages = []
    weaknesses = []
    for key in sorted(set(student_metrics) & set(baseline_metrics)):
        baseline = baseline_metrics[key]
        student = student_metrics[key]
        diff = student - baseline
        percent = (diff / baseline * 100.0) if baseline else None
        prefer_lower = _suite_prefers_lower(key, metadata)
        is_advantage = diff < 0 if prefer_lower else diff > 0
        row = {
            "metric": key,
            "baseline": round(baseline, 6),
            "student": round(student, 6),
            "diff": round(diff, 6),
            "diff_percent": round(percent, 2) if percent is not None else None,
            "judgement": "advantage" if is_advantage else "weakness",
        }
        comparisons.append(row)
        sentence = f"{key}: 学生结果 {student:.4g}，baseline {baseline:.4g}"
        if percent is not None:
            sentence += f"，差异 {percent:+.2f}%"
        if is_advantage:
            advantages.append(sentence)
        else:
            weaknesses.append(sentence)

    sample_count = (
        evaluation.get("sample_count")
        or evaluation.get("phase_count")
        or evaluation.get("valid_paths")
        or len(comparisons)
    )
    if comparisons:
        summary = (
            f"已完成 {task_name} 的任务专用评测，样本/相位数量 {sample_count}。"
            f"共识别 {len(comparisons)} 个可与 baseline 直接对比的指标，"
            f"其中 {len(advantages)} 项优于 baseline，{len(weaknesses)} 项弱于或未优于 baseline。"
        )
    elif student_metrics or comparison_metrics:
        summary = (
            f"已完成 {task_name} 的任务专用评测，结果文件格式可解析。"
            "该任务输出的是相对参考基线的评测指标，建议结合详细指标判断算法质量。"
        )
    else:
        summary = (
            f"已保存 {task_name} 提交文件，但任务评测器未返回可展示的数值指标。"
            "请检查结果文件是否使用基础数据包中的 student_submission_template 字段。"
        )

    advice = []
    if weaknesses:
        advice.append("优先检查弱势指标对应的约束、目标函数权重和数据样本覆盖情况。")
    if comparison_metrics:
        advice.append("本次评测已生成任务专用 comparison 指标，教师可在详情中查看完整 JSON。")
    advice.append("代码、方案说明和结果文件均已归档，教师端可按班级和学生下载复核。")

    text_parts = [summary]
    if advantages:
        text_parts.append("优势：" + "；".join(advantages[:5]) + "。")
    if weaknesses:
        text_parts.append("劣势/风险：" + "；".join(weaknesses[:5]) + "。")
    if comparison_metrics:
        text_parts.append(
            "评测补充：" + "；".join(f"{key}={value:.6g}" for key, value in list(comparison_metrics.items())[:8]) + "。"
        )
    text_parts.append("建议：" + "；".join(advice))
    baseline_file = metadata.get("baseline_file")
    if not baseline_file and (_PACK_ROOT / algorithm_type / "baseline_predictions.csv").is_file():
        baseline_file = "baseline_predictions.csv"

    return {
        "summary": summary,
        "advantages": advantages,
        "weaknesses": weaknesses,
        "advice": advice,
        "baseline_file": baseline_file,
        "primary_metric": metadata.get("primary_metric")
        or ((metadata.get("evaluation") if isinstance(metadata.get("evaluation"), dict) else {}) or {}).get("primary_metric"),
        "baseline_metrics": baseline_metrics,
        "student_metrics": student_metrics,
        "comparison_metrics": comparison_metrics,
        "comparisons": comparisons,
        "suite_evaluation": evaluation,
        "analysis_text": "\n".join(text_parts),
        "analysis_provider": "suite",
    }


def _call_ai_analysis(
    *,
    algorithm_type: str,
    metadata: dict[str, Any],
    builtin_analysis: dict[str, Any],
) -> str | None:
    settings = get_settings()
    api_url = (settings.self_algorithm_ai_api_url or "").strip()
    api_key = (settings.self_algorithm_ai_api_key or os.environ.get("OPENAI_API_KEY", "")).strip()
    if not api_url and api_key:
        api_url = "https://api.openai.com/v1/chat/completions"
    if not api_url:
        return None

    pack = _ALGORITHM_PACKS.get(algorithm_type, {})
    payload_for_model = {
        "algorithm_type": algorithm_type,
        "algorithm_label": pack.get("label", algorithm_type),
        "task_name": metadata.get("task_name"),
        "primary_metric": metadata.get("primary_metric"),
        "baseline_file": builtin_analysis.get("baseline_file"),
        "comparisons": builtin_analysis.get("comparisons", [])[:20],
        "baseline_metrics": _trim_dict(builtin_analysis.get("baseline_metrics", {})),
        "student_metrics": _trim_dict(builtin_analysis.get("student_metrics", {})),
        "builtin_summary": builtin_analysis.get("summary"),
        "builtin_advantages": builtin_analysis.get("advantages", [])[:8],
        "builtin_weaknesses": builtin_analysis.get("weaknesses", [])[:8],
    }
    system_prompt = (
        "你是面向车路云一体化综合交通实验平台的算法助教。"
        "请基于 baseline 与学生仿真结果的结构化对比，给出中文反馈。"
        "反馈必须包含：总体判断、相对 baseline 的优势、劣势或潜在风险、建议学生下一步检查的技术点。"
        "如果可比较指标不足，要明确指出输出格式问题，并建议按模板字段重新导出。"
        "不要编造未提供的指标，不要给出空泛鼓励。"
    )
    user_prompt = "请分析以下自研算法提交：\n" + json.dumps(payload_for_model, ensure_ascii=False, indent=2)
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    body = {
        "model": settings.self_algorithm_ai_model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.2,
    }
    try:
        with httpx.Client(timeout=settings.self_algorithm_ai_timeout_seconds) as client:
            resp = client.post(api_url, headers=headers, json=body)
        resp.raise_for_status()
        data = resp.json()
    except Exception:
        return None

    if isinstance(data, dict):
        direct = data.get("analysis_text") or data.get("text") or data.get("output_text")
        if isinstance(direct, str) and direct.strip():
            return direct.strip()
        choices = data.get("choices")
        if isinstance(choices, list) and choices:
            first = choices[0]
            if isinstance(first, dict):
                message = first.get("message")
                if isinstance(message, dict):
                    content = message.get("content")
                    if isinstance(content, str) and content.strip():
                        return content.strip()
    return None


def _attach_ai_analysis(
    *,
    algorithm_type: str,
    metadata: dict[str, Any],
    analysis: dict[str, Any],
) -> dict[str, Any]:
    ai_text = _call_ai_analysis(
        algorithm_type=algorithm_type,
        metadata=metadata,
        builtin_analysis=analysis,
    )
    if ai_text:
        analysis["ai_analysis_text"] = ai_text
        analysis["analysis_text"] = ai_text + "\n\n规则分析参考：\n" + analysis["analysis_text"]
        analysis["analysis_provider"] = "ai"
    return analysis


def _build_analysis(algorithm_type: str, result_path: Path) -> dict[str, Any]:
    item = _require_pack(algorithm_type)
    pack_dir = _PACK_ROOT / item["dir"]
    metadata = _read_metadata(pack_dir)
    suite_evaluation, suite_error = _run_suite_evaluation(algorithm_type, result_path)
    if suite_evaluation:
        return _attach_ai_analysis(
            algorithm_type=algorithm_type,
            metadata=metadata,
            analysis=_analysis_from_suite_evaluation(algorithm_type, metadata, suite_evaluation),
        )

    baseline_file = metadata.get("baseline_file")
    if not baseline_file and (pack_dir / "baseline_predictions.csv").is_file():
        baseline_file = "baseline_predictions.csv"
    baseline_path = pack_dir / baseline_file if baseline_file else None
    baseline_metrics = _extract_metrics(baseline_path) if baseline_path and baseline_path.is_file() else {}
    student_metrics = _extract_metrics(result_path)

    comparisons = []
    advantages = []
    weaknesses = []
    for key in sorted(set(baseline_metrics) & set(student_metrics)):
        baseline = baseline_metrics[key]
        student = student_metrics[key]
        diff = student - baseline
        percent = (diff / baseline * 100.0) if baseline else None
        prefer_lower = _metric_prefers_lower(key)
        is_advantage = diff < 0 if prefer_lower else diff > 0
        row = {
            "metric": key,
            "baseline": round(baseline, 6),
            "student": round(student, 6),
            "diff": round(diff, 6),
            "diff_percent": round(percent, 2) if percent is not None else None,
            "judgement": "advantage" if is_advantage else "weakness",
        }
        comparisons.append(row)
        sentence = f"{key}: 学生结果 {student:.4g}，baseline {baseline:.4g}"
        if percent is not None:
            sentence += f"，差异 {percent:+.2f}%"
        if is_advantage:
            advantages.append(sentence)
        else:
            weaknesses.append(sentence)

    if comparisons:
        summary = (
            f"已完成与 {baseline_file} 的智能对比。"
            f"共识别 {len(comparisons)} 个可比较数值指标，"
            f"其中 {len(advantages)} 项优于 baseline，{len(weaknesses)} 项弱于或未优于 baseline。"
        )
    else:
        summary = (
            "已保存提交文件，但结果文件未解析出与 baseline 同名的数值指标。"
            "建议按照基础数据包中的 student_submission_template 字段输出 CSV/JSON/XLSX，"
            "便于平台自动给出更精确的优势和劣势分析。"
        )
    if suite_error:
        summary += f"\n任务专用评测提示：{suite_error}"

    advice = []
    if weaknesses:
        advice.append("优先检查弱势指标对应的约束、目标函数权重和仿真随机种子，确认与 baseline 使用同一场景。")
    if not comparisons:
        advice.append("结果文件建议至少包含模板中的关键字段，并保留表头或 JSON 键名。")
    if suite_error:
        advice.append("可直接复用下载包中的 student_submission_template 文件头重新导出结果。")
    advice.append("教师端可按机构-班级-学生查看本次提交，代码、方案说明和结果文件均已归档。")

    text_parts = [summary]
    if advantages:
        text_parts.append("优势：" + "；".join(advantages[:5]) + "。")
    if weaknesses:
        text_parts.append("劣势/风险：" + "；".join(weaknesses[:5]) + "。")
    text_parts.append("建议：" + "；".join(advice))

    analysis = {
        "summary": summary,
        "advantages": advantages,
        "weaknesses": weaknesses,
        "advice": advice,
        "baseline_file": baseline_file,
        "primary_metric": metadata.get("primary_metric"),
        "baseline_metrics": baseline_metrics,
        "student_metrics": student_metrics,
        "comparisons": comparisons,
        "analysis_text": "\n".join(text_parts),
        "analysis_provider": "builtin",
    }
    if suite_error:
        analysis["suite_evaluation_error"] = suite_error
    return _attach_ai_analysis(algorithm_type=algorithm_type, metadata=metadata, analysis=analysis)


def _apply_access_filter(q, current_user: User):
    if current_user.role == "student":
        return q.filter(SelfAlgorithmSubmission.student_id == current_user.id)
    if current_user.role == "teacher":
        return q.join(Class, SelfAlgorithmSubmission.class_id == Class.id).filter(Class.teacher_id == current_user.id)
    if current_user.role == "org_admin":
        return q.filter(SelfAlgorithmSubmission.org_id == current_user.org_id)
    return q


def _get_accessible_submission(db: Session, submission_id: int, current_user: User) -> SelfAlgorithmSubmission:
    q = db.query(SelfAlgorithmSubmission).filter(SelfAlgorithmSubmission.id == submission_id)
    item = _apply_access_filter(q, current_user).first()
    if item is None:
        raise HTTPException(status_code=404, detail="自研算法提交记录不存在")
    return item


@router.get("/packs", response_model=list[SelfAlgorithmPack])
def list_packs(current_user: User = Depends(get_current_user)):
    return [
        _pack_to_schema(algorithm_type, item)
        for algorithm_type, item in _ALGORITHM_PACKS.items()
        if (_PACK_ROOT / item["dir"]).is_dir()
    ]


@router.get("/packs/{algorithm_type}/download")
def download_pack(
    algorithm_type: str,
    current_user: User = Depends(require_role("student", "teacher", "org_admin", "super_admin")),
):
    item = _require_pack(algorithm_type)
    pack_dir = _PACK_ROOT / item["dir"]
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in sorted(pack_dir.rglob("*")):
            if file_path.is_file():
                zf.write(file_path, arcname=str(file_path.relative_to(pack_dir)))
    buffer.seek(0)
    filename = f"{algorithm_type}_基础数据包.zip"
    headers = _download_headers(filename, f"{algorithm_type}_pack.zip")
    return StreamingResponse(buffer, media_type="application/zip", headers=headers)


@router.post("/submissions", response_model=SelfAlgorithmSubmissionRead, status_code=status.HTTP_201_CREATED)
def submit_self_algorithm(
    algorithm_type: str = Form(...),
    class_id: int | None = Form(None),
    code_file: UploadFile = File(...),
    spec_file: UploadFile = File(...),
    result_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("student")),
):
    _require_pack(algorithm_type)
    code_name, code_content = _validate_upload(code_file, _CODE_EXTS, "代码")
    spec_name, spec_content = _validate_upload(spec_file, _SPEC_EXTS, "方案说明")
    result_name, result_content = _validate_upload(result_file, _RESULT_EXTS, "仿真结果")

    target_class_id = _get_student_class_id(db, current_user, class_id)
    cls = db.query(Class).filter(Class.id == target_class_id).first()
    if cls is None:
        raise HTTPException(status_code=404, detail="班级不存在")
    org_id = cls.org_id or current_user.org_id

    settings = get_settings()
    folder = (
        Path(settings.upload_dir)
        / "self_algorithms"
        / f"org_{org_id or 'none'}"
        / f"class_{target_class_id}"
        / f"student_{current_user.id}"
        / f"{algorithm_type}_{uuid.uuid4().hex[:12]}"
    )
    code_path = _save_file(folder, "01_code_" + code_name, code_content)
    spec_path = _save_file(folder, "02_spec_" + spec_name, spec_content)
    result_path = _save_file(folder, "03_result_" + result_name, result_content)

    analysis = _build_analysis(algorithm_type, Path(result_path))
    analysis_text = analysis.pop("analysis_text")
    analysis_provider = analysis.pop("analysis_provider", "builtin")

    submission = SelfAlgorithmSubmission(
        algorithm_type=algorithm_type,
        org_id=org_id,
        class_id=target_class_id,
        student_id=current_user.id,
        code_file_path=code_path,
        code_original_filename=code_name,
        code_file_size=len(code_content),
        spec_file_path=spec_path,
        spec_original_filename=spec_name,
        spec_file_size=len(spec_content),
        result_file_path=result_path,
        result_original_filename=result_name,
        result_file_size=len(result_content),
        analysis_status="done",
        analysis_provider=analysis_provider,
        analysis_text=analysis_text,
        analysis_json=json.dumps(analysis, ensure_ascii=False),
    )
    db.add(submission)
    db.commit()
    db.refresh(submission)
    return submission


@router.get("/submissions", response_model=PagedResponse[SelfAlgorithmSubmissionRead])
def list_submissions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    algorithm_type: str | None = Query(None),
    class_id: int | None = Query(None),
    student_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = _apply_access_filter(db.query(SelfAlgorithmSubmission), current_user)
    if algorithm_type:
        q = q.filter(SelfAlgorithmSubmission.algorithm_type == algorithm_type)
    if class_id and current_user.role != "student":
        q = q.filter(SelfAlgorithmSubmission.class_id == class_id)
    if student_id and current_user.role != "student":
        q = q.filter(SelfAlgorithmSubmission.student_id == student_id)
    total = q.count()
    pages = (total + page_size - 1) // page_size if total > 0 else 1
    items = q.order_by(SelfAlgorithmSubmission.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return PagedResponse(items=items, total=total, page=page, page_size=page_size, pages=pages)


@router.get("/submissions/{submission_id}/download/{kind}")
def download_submission_file(
    submission_id: int,
    kind: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    submission = _get_accessible_submission(db, submission_id, current_user)
    mapping = {
        "code": (submission.code_file_path, submission.code_original_filename),
        "spec": (submission.spec_file_path, submission.spec_original_filename),
        "result": (submission.result_file_path, submission.result_original_filename),
    }
    if kind not in mapping:
        raise HTTPException(status_code=400, detail="下载类型必须是 code/spec/result")
    file_path, filename = mapping[kind]
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(file_path, filename=filename or Path(file_path).name)
